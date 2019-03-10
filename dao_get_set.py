#coding=gbk
import openpyxl,os,sys
def is_cofig_exists():
    is_exists = os.path.exists(os.path.join(os.getcwd() ,'config.ini'))
    if not is_exists:
        print('Error: config.ini is not found')
        return False
    return True

def read_excel_name():
    if not is_cofig_exists():
        return ''
    config_path = os.path.join(os.getcwd() ,'config.ini')
    config_file = open(config_path, 'r')
    excel_name = config_file.readline()
    if excel_name == '' :
        print('Error: the name of excel is not found. please fill in your name of excel in the first line')
        return ''
    return excel_name

def is_excel_exists():
    excel_name = read_excel_name()
    if excel_name == '' :
        return False
    excel_path = os.path.join(os.getcwd(), excel_name).replace('\n','').strip()
    is_excel_path_exists = os.path.exists(excel_path)
    if not is_excel_path_exists :
        print('Error: the excel is not found. Please chick the excel is if exists')
        return False
    return True;

def get_excel_name():
    if not is_excel_exists() :
        return ''
    excel_name = read_excel_name()
    return excel_name

def creat_result_path():
    result_file_dir = os.path.join(os.getcwd() + os.path.sep + 'result')
    isExists=os.path.exists(result_file_dir)
    if not isExists :
        os.makedirs(result_file_dir)
    return result_file_dir;

    
def fun():
    excel_name = get_excel_name()
    if excel_name == '' :
        return False
    sheet = openpyxl.load_workbook(excel_name).get_sheet_by_name('Sheet1')

    object_name = sheet['A2'].value
    if object_name == None:
        object_name = 'uPig'
    result_name = sheet['B2'].value
    if result_name == None:
        result_name = 'uDog'
    list_name = sheet['C2'].value
    
    maxRow = sheet.max_row
    
    set_result = ''
    for i in range(5, maxRow):
        property_name = sheet.cell(row = i, column = 1).value
        property_type = sheet.cell(row = i, column = 2).value
        index = sheet.cell(row = i, column = 3).value
        get_result = result_name + '.get' + property_type.capitalize() + '(' +str(index) + ')'
        set_result += object_name + '.set' + property_name.capitalize() + '(' + get_result + ');\n'
    
    if not list_name == None :
        list_add_str = list_name + '.add(' + object_name + ');\n'
        set_result += list_add_str
    print(set_result)
    
    result_file_dir = creat_result_path()
    current_file_path = os.path.join(result_file_dir, excel_name.replace('.xlsx', '').strip() + '.java')
    
    file = open(current_file_path, 'w')
    
    file.write(set_result)
    
    file.close
    
    return True

def main():
    fun()
    os.system('pause')
    
if __name__ == '__main__':
    sys.exit(main())
    
